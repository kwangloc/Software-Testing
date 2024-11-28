from selenium import webdriver
from selenium.webdriver.chrome.service import Service # set up chrome driver service 
import unittest # built-in python lib
import page
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


class TestingPHPTravels(unittest.TestCase):
    homepage_url = 'https://phptravels.net/'

    # File signup
    test_signup_input_file = 'test_case.xlsx'  
    # test_signup_output_file = 'test_signup_result.xlsx'
    test_signup_output_file = 'test_signup_result.xlsx'
    # sheet_signup = 'test_case_signup'
    sheet_signup = 'temp_signup'
    columns_signup = ['Test case', 'Email', 'Password', 'Firstname', 'Lastname', 'Country', 'Phone', 'Describe', 'Expected output']
    test_signup_results = []

    # File login
    # test_login_input_file = 'test_case.xlsx'  
    # sheet_login = 'test_case_login'
    # test_login_output_file = 'test_login_result.xlsx'
    # columns_login = ['Test case', 'Email', 'Password', 'Describe', 'Expected output']

    def setUp(self):
        print('Setting up.')
        self.service = Service(executable_path='../chromedriver.exe')
        self.driver = webdriver.Chrome(service=self.service)
        # self.driver.get(self.homepage_url)
        self.data = pd.read_excel(self.test_signup_input_file, sheet_name=self.sheet_signup)

    def test_signup(self):
        for index, row in self.data.iterrows():
            print('---------Describe:', row['Describe'])
            # user = {col: row[col] for col in self.columns_signup if pd.notna(row[col])}
            user = {col: (row[col] if pd.notna(row[col]) else "") for col in self.columns_signup}

            try:
                # Go to main page and check the title
                self.driver.get(self.homepage_url)
                mainPage = page.MainPage(self.driver)
                assert mainPage.does_title_match(), "Homepage title does not match."

                # Navigate to signup page and check the title
                mainPage.go_signup_page()
                signup_page = page.SignupPage(self.driver)
                assert signup_page.does_title_match(), "Signup page title does not match."

                # Fill and submit signup form
                # signup_page.fill_n_submit(user)
                time.sleep(2)
                submission_success = signup_page.fill_n_submit(user)
                # time.sleep(10)
                if not submission_success:
                    self.test_signup_results.append('Failed')
                    print('Failed at not submission_success')
                    continue  # Skip further checks if submission did not occur
                # 
                time.sleep(2)
                # test_signup_results.append('passed')
                # assert signup_page.is_signup_success()
                #
                if signup_page.is_signup_success():
                    self.test_signup_results.append('Passed')
                    print('Passed')
                else:
                    self.test_signup_results.append('Failed')
                    print('Failed at is_signup_success')

            except Exception as err:
                print(f"-> Test case {row['Test case']} failed with error: {err}")
                self.test_signup_results.append('Failed')
                print('Failed at Exception')

                # pass
                # assert False
        print(self.test_signup_results)
        self.data['Actual result'] = self.test_signup_results
        print(self.data)
        write_test_results_from_dataframe(df=self.data, output_file=self.test_signup_output_file, start_row=11)

    def tearDown(self):
        self.driver.quit()

def suite():
    suite = unittest.TestSuite()
    suite.addTest(TestingPHPTravels('test_signup'))
    return suite

def write_test_results_from_dataframe(df, output_file, start_row=11):
    # Ensure the DataFrame includes the "Actual result" column
    if "Actual result" not in df.columns:
        raise ValueError("The DataFrame must include an 'Actual result' column.")

    # Load the existing workbook to preserve existing content
    wb = load_workbook(output_file)
    ws = wb.active

    # Write headers to the specified start row
    for col_num, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    # Write data starting from the next row
    for row_idx, row in enumerate(df.values, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    # Save the file to preserve both default description and new data
    wb.save(output_file)
    print(f"Test results written successfully to {output_file}")

if __name__ == '__main__':
    # unittest.main()
    runner = unittest.TextTestRunner()
    runner.run(suite())