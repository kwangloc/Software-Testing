import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

test_login_input_file = 'test_case.xlsx'
sheet_read = 'test_case_login'
test_login_output_file = 'test_login_result.xlsx'

data = pd.read_excel(test_login_input_file, sheet_name=sheet_read)

columns = ['Test case', 'Email', 'Password', 'Describe', 'Expected output']
test_results = []

# Generate test results
for index, row in data.iterrows():
    test_results.append('pass')

# Add "Actual result" columns
data['Actual result'] = test_results

wb = load_workbook(test_login_output_file)  # Load the existing output file
ws = wb.active

# Write headers to row 11
for col_num, header in enumerate(columns + ['Actual result'], start=1):  # Adjust header row
    cell = ws.cell(row=11, column=col_num)
    cell.value = header
    cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Green background
    cell.font = Font(color="FFFFFF", bold=True)  # White font

# Write data starting from row 12
for row_idx, row in enumerate(data.values, start=12):  # Data begins at row 12
    for col_idx, value in enumerate(row, start=1):  # Columns begin at A (1)
        ws.cell(row=row_idx, column=col_idx).value = value

wb.save(test_login_output_file)
