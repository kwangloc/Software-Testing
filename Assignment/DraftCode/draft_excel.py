import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def write_test_results_from_dataframe(df, output_file, start_row=11):
    # Ensure the DataFrame includes the "Actual result" column
    if "Actual result" not in df.columns:
        raise ValueError("The DataFrame must include an 'Actual result' column.")

    # Load the existing workbook to preserve existing content
    wb = load_workbook(output_file)
    ws = wb.active

    # Write headers to the specified start row
    for col_num, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    # Write data starting from the next row
    for row_idx, row in enumerate(df.values, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    # Save the file to preserve both default description and new data
    wb.save(output_file)
    print(f"Test results written successfully to {output_file}")

# Example usage

test_signup_input_file = 'test_case.xlsx'  
test_signup_output_file = 'test_signup_result_3.xlsx'
sheet_signup = 'temp_signup'
columns_signup = ['Test case', 'Email', 'Password', 'Firstname', 'Lastname', 'Country', 'Phone', 'Describe', 'Expected output']
test_signup_results = []
data = pd.read_excel(test_signup_input_file, sheet_name=sheet_signup)

for index, row in data.iterrows():
    test_signup_results.append('AWAWAW')
data['Actual result'] = test_signup_results

write_test_results_from_dataframe(
    df=data,
    output_file=test_signup_output_file,
    start_row=11
)
